from io import BytesIO
from os import getenv

from dotenv import load_dotenv
from flask import Flask, jsonify, request, Response
from openpyxl.reader.excel import load_workbook
from openpyxl.styles.borders import Border

load_dotenv()

app = Flask(__name__)
allowed_addresses = {"127.0.0.1", "localhost", "::1"}


@app.before_request
def limit_remote_addr() -> tuple[Response, int] | None:
    if request.remote_addr not in allowed_addresses:
        return jsonify({"error": "Access denied."}), 403


@app.route("/api/parse-xlsx-borders", methods=["POST"])
def parse_xlsx_borders() -> tuple[Response, int]:
    if "file" not in request.files:
        return jsonify({"error": "No file provided."}), 400

    file = request.files["file"]

    if file.mimetype != "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet":
        return jsonify({"error": "Invalid file format. Expected XLSX file."}), 400

    file_content = BytesIO(file.read())
    try:
        borders = get_xlsx_borders(file_content)
        return jsonify(borders), 200
    except Exception as e:
        return jsonify({"error": f"Error processing file: {e}"}), 500


def get_xlsx_borders(buffer: BytesIO) -> list[list[dict[str, bool]]]:
    workbook = load_workbook(buffer)
    all_borders: list[list[dict[str, bool]]] = []

    for worksheet in workbook.worksheets:
        for cells_row in worksheet.iter_rows():
            borders_row: list[dict[str, bool]] = []

            for cell in cells_row:
                # noinspection PyTypeChecker
                cell_border: Border = cell.border
                has_top = cell_border.top is not None and cell_border.top.style is not None
                has_bottom = cell_border.bottom is not None and cell_border.bottom.style is not None
                has_left = cell_border.left is not None and cell_border.left.style is not None
                has_right = cell_border.right is not None and cell_border.right.style is not None

                borders_row.append({
                    "top": has_top,
                    "bottom": has_bottom,
                    "left": has_left,
                    "right": has_right,
                })

            all_borders.append(borders_row)

    return all_borders


if __name__ == "__main__":
    app.run(host="127.0.0.1", port=getenv("PY_PORT"))
